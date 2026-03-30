"""
Gestión de Deuda Municipal — Streamlit App
Municipalidad de General La Madrid · R.A.F.A.M.

Instalación:
    pip install -r requirements.txt

Ejecución:
    streamlit run app.py
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import io
import csv
import os
from datetime import datetime, date
from typing import Optional

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
    _PDF_AZUL = colors.HexColor("#2B3D6B")
    _PDF_AZUL_CLARO = colors.HexColor("#4A6FA5")
    HAS_REPORTLAB = True
except ImportError:
    _PDF_AZUL = None
    _PDF_AZUL_CLARO = None
    HAS_REPORTLAB = False

# ─── CONSTANTES (antes de set_page_config para favicon) ────────────────────────
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
_ASSETS_LOCAL = os.path.join(_BASE_DIR, "assets")
_LOGO_FALLBACKS = [
    os.path.join(_ASSETS_LOCAL, "logo_municipio.png"),
    os.path.join(_ASSETS_LOCAL, "logo_ip.png"),
    r"C:\Users\lazap\.cursor\projects\c-Users-lazap-OneDrive-Escritorio\assets\c__Users_lazap_AppData_Roaming_Cursor_User_workspaceStorage_8a7a64bc412e995d63fda3f331415237_images_LOGO_MUNICIPIO__3_-8244affc-de13-4319-afd7-d332d1800c09.png",
    r"C:\Users\lazap\.cursor\projects\c-Users-lazap-OneDrive-Escritorio\assets\c__Users_lazap_AppData_Roaming_Cursor_User_workspaceStorage_8a7a64bc412e995d63fda3f331415237_images_SIMBOLO_BLANCO_FNDO_AZUL-1cd3534f-1bbc-4bc9-b592-caa615000a90.png",
    r"C:\Users\lazap\.cursor\projects\c-Users-lazap-OneDrive-Escritorio\assets\c__Users_lazap_AppData_Roaming_Cursor_User_workspaceStorage_8a7a64bc412e995d63fda3f331415237_images_SIMBOLO_BLANCO_FNDO_AZUL-d3d796d7-9c3a-43b5-b819-88de7b100ba2.png",
]

def _first_existing(paths):
    for p in paths:
        if p and os.path.isfile(p):
            return p
    return None

FAVICON_PATH = _first_existing([os.path.join(_ASSETS_LOCAL, "logo_ip.png"), _LOGO_FALLBACKS[3], _LOGO_FALLBACKS[4]])

st.set_page_config(
    page_title="Gestión de Deuda · Municipalidad GLM",
    page_icon=FAVICON_PATH if FAVICON_PATH else "🏛️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CUSTOM CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,500&family=IBM+Plex+Sans:wght@300;400;500;600;700&family=Source+Sans+3:ital,wght@0,400;0,600;0,700;1,400&display=swap');

:root {
    --azul:      #1E3A5F;
    --azul-med:  #2E6DA4;
    --azul-clr:  #D6E4F0;
    --verde:     #1A6B3C;
    --rojo:      #C0392B;
    --naranja:   #D68910;
    --gris-bg:   #F0F4F8;
    --gris-bdr:  #CBD5E1;
    --blanco:    #FFFFFF;
    --texto:     #0F172A;
    --texto2:    #475569;
    --fondo:     #EEF2F7;
}

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif !important;
}

.stApp { background-color: var(--fondo); }
.main .block-container { padding: 1.5rem 2rem; max-width: 100%; }

[data-testid="stSidebar"] {
    background: var(--azul) !important;
    border-right: 3px solid #13273F;
}
[data-testid="stSidebar"] * { color: #CBD5E1 !important; }

.kpi-card {
    background: var(--blanco);
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    border-left: 5px solid var(--azul);
    box-shadow: 0 1px 6px rgba(30,58,95,0.10);
    transition: transform 0.15s;
    margin-bottom: 0.5rem;
    min-height: 126px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
}
.kpi-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(30,58,95,0.15); }
.kpi-card.verde   { border-color: var(--verde); }
.kpi-card.rojo    { border-color: var(--rojo); }
.kpi-card.naranja { border-color: var(--naranja); }
.kpi-label { font-size: 0.72rem; font-weight: 600; text-transform: uppercase;
             letter-spacing: 0.08em; color: var(--texto2); margin-bottom: 0.25rem; }
.kpi-value { font-size: clamp(1.25rem, 2.2vw, 1.85rem); font-weight: 700; color: var(--azul);
             font-family: 'DM Sans', 'Source Sans 3', 'IBM Plex Sans', sans-serif; letter-spacing: -0.01em;
             font-variant-numeric: tabular-nums;
             line-height: 1.15;
             white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.kpi-value.verde   { color: var(--verde); }
.kpi-value.rojo    { color: var(--rojo); }
.kpi-value.naranja { color: var(--naranja); }
.kpi-sub { font-size: 0.75rem; color: var(--texto2); margin-top: 0.2rem; }

.sec-header {
    font-size: 0.8rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.1em; color: var(--azul);
    border-bottom: 2px solid var(--azul-clr);
    padding-bottom: 0.4rem; margin: 1.2rem 0 0.8rem;
}

.page-banner {
    background: linear-gradient(135deg, var(--azul) 0%, #2A4F7C 100%);
    color: white; border-radius: 10px; padding: 1.2rem 1.8rem;
    margin-bottom: 1.5rem;
    box-shadow: 0 3px 15px rgba(30,58,95,0.25);
}
.page-banner h2 { font-size: 1.4rem; font-weight: 700; margin: 0 0 0.2rem; color: white !important; }
.page-banner p  { font-size: 0.8rem; color: #90B8D8; margin: 0; }

.info-box {
    background: var(--azul-clr); border-left: 4px solid var(--azul-med);
    padding: 0.7rem 1rem; border-radius: 0 6px 6px 0;
    font-size: 0.85rem; color: var(--azul); margin-bottom: 1rem;
}
.warn-box {
    background: #FEF9E7; border-left: 4px solid var(--naranja);
    padding: 0.7rem 1rem; border-radius: 0 6px 6px 0;
    font-size: 0.85rem; color: #7D5300; margin-bottom: 1rem;
}

.stButton > button {
    background: var(--azul) !important; color: white !important;
    border: none; border-radius: 6px; font-weight: 600;
    transition: background 0.2s;
}
.stButton > button:hover { background: #2A4F7C !important; }
[data-testid="stSidebar"] .stButton > button {
    background: rgba(214, 228, 240, 0.14) !important;
    border: 1px solid rgba(214, 228, 240, 0.35) !important;
    color: #E8F0FB !important;
    border-radius: 8px !important;
    backdrop-filter: blur(1px);
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: rgba(214, 228, 240, 0.22) !important;
    border-color: rgba(214, 228, 240, 0.55) !important;
}

[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }

.filter-panel {
    background: rgba(255, 255, 255, 0.55);
    border: 1px solid rgba(30, 58, 95, 0.12);
    border-radius: 10px;
    padding: 0.85rem 1rem 0.5rem;
    margin-bottom: 1rem;
    box-shadow: 0 1px 3px rgba(30, 58, 95, 0.06);
}
/* Inputs más visibles */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {
    background: rgba(30, 58, 95, 0.06) !important;
    border: 1px solid rgba(30, 58, 95, 0.22) !important;
}
[data-testid="stSelectbox"] [data-baseweb="select"] > div {
    background: rgba(30, 58, 95, 0.06) !important;
    border: 1px solid rgba(30, 58, 95, 0.22) !important;
}
/* Spinner de carga: rueda solo borde, centrada */
@keyframes spin-border {
    to { transform: rotate(360deg); }
}
.loading-wheel {
    width: 48px; height: 48px;
    border: 3px solid var(--azul-clr);
    border-top-color: var(--azul);
    border-radius: 50%;
    animation: spin-border 0.8s linear infinite;
}
[data-testid="stSpinner"] > div {
    border: 3px solid var(--azul-clr) !important;
    border-top-color: var(--azul) !important;
    border-radius: 50% !important;
    width: 48px !important;
    height: 48px !important;
    animation: spin-border 0.8s linear infinite !important;
    background: transparent !important;
}

/* Contenedores con borde (Streamlit border=True) más visibles */
div[data-testid="stVerticalBlockBorderWrapper"] {
    border: 1.5px solid rgba(30, 58, 95, 0.35) !important;
    background: rgba(255, 255, 255, 0.78) !important;
    border-radius: 10px !important;
    box-shadow: 0 2px 8px rgba(30, 58, 95, 0.07) !important;
    padding: 0.4rem 0.55rem 0.25rem !important;
}
.mini-kpi {
    background: rgba(255, 255, 255, 0.78);
    border: 1px solid rgba(30, 58, 95, 0.16);
    border-left: 3px solid rgba(30, 58, 95, 0.45);
    border-radius: 8px;
    padding: 0.55rem 0.65rem;
    min-height: 66px;
}
.mini-kpi .label {
    font-size: 0.68rem;
    color: #475569;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    font-weight: 600;
}
.mini-kpi .value {
    margin-top: 0.15rem;
    font-size: 1rem;
    font-weight: 700;
    color: #1E3A5F;
}
</style>
""", unsafe_allow_html=True)

# ─── CONSTANTES (resto) ───────────────────────────────────────────────────────
GESTION_COLS = [
    "CUIM", "APELLIDO Y NOMBRE", "NRO. IMPONIBLE", "RECURSO",
    "TIPO IMPONIBLE", "CAPITAL", "INTERÉS", "DEUDA TOTAL",
    "TELÉFONO", "MAIL", "OBSERVACIONES",
    "NOTIF. LEGAL", "NOTIF. CÉDULA", "CARTA DOC.",
    "ESTADO GESTIÓN",
]
ESTADOS = ["Pendiente", "En Gestión", "Legales", "Regularizado"]
YES_VALUES = {"Sí", "Si", "si", "sí", "1", "True", "true", "SÍ", "SI"}
NUMERIC_COLS_BASE = ["CAPITAL", "INTERÉS", "DEUDA TOTAL", "MULTA"]
DATE_NOTIF_COLS = ["NOTIF. LEGAL", "NOTIF. CÉDULA", "CARTA DOC."]
STATE_FILE = "gestion_deuda_state.pkl"
LOGO_PATH = "logo.png"
LOGO_IP_PATH = "favicon.png"

OFICINA_MEMBRETE = (
    "Oficina de Ingresos Públicos\n"
    "Dirección San Martín 565 — General La Madrid, Buenos Aires\n"
    "ingresospublicos@lamadrid.gob.ar · 2284534618"
)


# ─── SESSION STATE ────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "df_datos":    pd.DataFrame(),
        "df_gestion":  pd.DataFrame(columns=GESTION_COLS),
        "page":        "dashboard",
        "last_update": None,
        "last_file_load": None,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

    load_persistent_state()

# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fmt_peso(v):
    try:
        return f"$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "$ 0,00"


def gestion_row_key(df: pd.DataFrame) -> pd.Series:
    """Clave única por línea: imponible + recurso."""
    n = df["NRO. IMPONIBLE"].astype(str).str.strip()
    r = (
        df["RECURSO"].astype(str).str.strip()
        if "RECURSO" in df.columns
        else pd.Series([""] * len(df), index=df.index)
    )
    return n + "|" + r


def notif_tiene_fecha(val) -> bool:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    try:
        return pd.notna(pd.to_datetime(val, errors="coerce"))
    except Exception:
        return False


def migrate_legacy_gestion(df: pd.DataFrame) -> pd.DataFrame:
    """Convierte pickles/planillas viejas al esquema actual sin perder datos manuales."""
    if df is None or df.empty:
        return df
    out = df.copy()
    for old in ["PRIORIDAD", "RESPONSABLE", "ÚLTIMA ACCIÓN", "FECHA NOTIF."]:
        if old in out.columns:
            out = out.drop(columns=[old], errors="ignore")
    for col in DATE_NOTIF_COLS:
        if col not in out.columns:
            continue
        new_vals = []
        for v in out[col].tolist():
            if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() == "":
                new_vals.append(pd.NaT)
            elif str(v).strip() in YES_VALUES:
                new_vals.append(pd.NaT)
            elif str(v).strip().lower() in ("no", "n"):
                new_vals.append(pd.NaT)
            else:
                new_vals.append(pd.to_datetime(v, errors="coerce"))
        out[col] = new_vals
    if "ESTADO GESTIÓN" in out.columns:
        emap = {
            "En proceso": "En Gestión",
            "Juicio": "Legales",
            "Incobreable": "Legales",
            "Plan de pago": "En Gestión",
        }
        out["ESTADO GESTIÓN"] = out["ESTADO GESTIÓN"].astype(str).str.strip().replace(emap)
        out.loc[~out["ESTADO GESTIÓN"].isin(ESTADOS), "ESTADO GESTIÓN"] = "Pendiente"
    return out


def normalize_numeric_columns(df: pd.DataFrame, cols=None) -> pd.DataFrame:
    """Convert known numeric columns once, to avoid repeated page-level casting."""
    if df.empty:
        return df
    cols = cols or NUMERIC_COLS_BASE
    for col in cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def order_gestion_df(df: pd.DataFrame) -> pd.DataFrame:
    """Orden fijo de columnas (plantilla municipal), aunque el archivo venga distinto."""
    if df is None or df.empty:
        return pd.DataFrame(columns=GESTION_COLS)
    out = migrate_legacy_gestion(df.copy())
    for c in GESTION_COLS:
        if c not in out.columns:
            if c in NUMERIC_COLS_BASE:
                out[c] = 0
            elif c in DATE_NOTIF_COLS:
                out[c] = pd.NaT
            else:
                out[c] = ""
    for c in DATE_NOTIF_COLS:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce")
    if "ESTADO GESTIÓN" in out.columns:
        out["ESTADO GESTIÓN"] = out["ESTADO GESTIÓN"].apply(
            lambda x: x if str(x).strip() in ESTADOS else "Pendiente"
        )
    return out[GESTION_COLS]


def excel_bytes_gestion_profesional(
    df_g: Optional[pd.DataFrame] = None,
    df_datos: Optional[pd.DataFrame] = None,
    *,
    include_gestion: bool = True,
    include_datos: bool = True,
) -> bytes:
    """Excel con hojas opcionales GESTION / DATOS y encabezados estilo planilla (rojo oscuro / blanco)."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    wrote = False
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if include_gestion and df_g is not None and not df_g.empty:
            order_gestion_df(df_g).to_excel(writer, sheet_name="GESTION", index=False)
            wrote = True
        if include_datos and df_datos is not None and not df_datos.empty:
            df_datos.to_excel(writer, sheet_name="DATOS", index=False)
            wrote = True
    if not wrote:
        raise ValueError("No hay datos para exportar con las opciones elegidas.")
    buf.seek(0)
    wb = load_workbook(buf)
    fill = PatternFill("solid", fgColor="9B2335")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = align
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _pdf_draw_header_footer(canvas, doc):
    """Membrete, marca de agua, borde inferior y número de hoja en cada página."""
    w, h = A4
    canvas.saveState()
    canvas.setAuthor("Municipalidad de General La Madrid - Ingresos Públicos")
    canvas.setTitle("Informe de deuda")
    canvas.setSubject("Reporte institucional de deuda")
    # Marca de agua: logo centrado, baja opacidad
    if LOGO_IP_PATH and os.path.isfile(LOGO_IP_PATH):
        try:
            canvas.saveState()
            canvas.setFillAlpha(0.06)
            canvas.drawImage(
                ImageReader(LOGO_IP_PATH),
                (w - 5 * cm) / 2,
                (h - 5 * cm) / 2,
                width=5 * cm,
                height=5 * cm,
                preserveAspectRatio=True,
                mask="auto",
            )
            canvas.restoreState()
        except Exception:
            pass
    # Header: logo municipal + logo IP + membrete
    if LOGO_PATH and os.path.isfile(LOGO_PATH):
        try:
            canvas.drawImage(
                ImageReader(LOGO_PATH),
                1.5 * cm,
                h - 2.35 * cm,
                width=1.5 * cm,
                height=1.5 * cm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass
    if LOGO_IP_PATH and os.path.isfile(LOGO_IP_PATH):
        try:
            canvas.drawImage(
                ImageReader(LOGO_IP_PATH),
                3.2 * cm,
                h - 2.15 * cm,
                width=1.35 * cm,
                height=1.35 * cm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass
    canvas.setFillColor(_PDF_AZUL)
    canvas.setFont("Helvetica-Bold", 8.5)
    y = h - 1.05 * cm
    for line in OFICINA_MEMBRETE.split("\n"):
        canvas.drawRightString(w - 1.5 * cm, y, line.strip())
        y -= 11
    # Footer: borde inferior institucional + número de hoja
    canvas.setStrokeColor(_PDF_AZUL)
    canvas.setLineWidth(4)
    canvas.line(0, 1.2 * cm, w, 1.2 * cm)
    canvas.setFillColor(_PDF_AZUL)
    canvas.setFont("Helvetica", 9)
    try:
        num = canvas.getPageNumber()
    except Exception:
        num = 1
    canvas.drawCentredString(w / 2, 0.5 * cm, f"Hoja {num}")
    canvas.restoreState()


def pdf_bytes_reporte_deuda(df_rows: pd.DataFrame, subtitulo: str) -> bytes:
    """PDF institucional con membrete en cada página."""
    if not HAS_REPORTLAB:
        raise RuntimeError("Instalá reportlab: pip install reportlab")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.4 * cm,
        rightMargin=1.4 * cm,
        topMargin=2.6 * cm,
        bottomMargin=2.2 * cm,
    )
    styles = getSampleStyleSheet()
    cell_st = ParagraphStyle(
        "cell",
        parent=styles["Normal"],
        fontSize=7,
        leading=8,
        wordWrap="CJK",
        leftIndent=0,
        rightIndent=0,
    )
    title_st = ParagraphStyle(
        "t",
        parent=styles["Heading1"],
        fontSize=13,
        textColor=_PDF_AZUL,
        spaceAfter=8,
        alignment=1,
    )
    sub_st = ParagraphStyle(
        "s",
        parent=styles["Normal"],
        fontSize=10,
        textColor=_PDF_AZUL_CLARO,
        spaceAfter=10,
        alignment=1,
    )
    intro_st = ParagraphStyle(
        "i",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#334155"),
        spaceAfter=14,
    )
    story = [
        Paragraph("Municipalidad de General La Madrid", title_st),
        Paragraph(f"<i>{subtitulo}</i>", sub_st),
        Paragraph(
            "Informe de situación de deuda con discriminación de capital e intereses, "
            "emitido para uso institucional de la Dirección de Ingresos Públicos.",
            intro_st,
        ),
    ]

    tw = doc.width
    col_w = [tw * 0.11, tw * 0.28, tw * 0.18, tw * 0.14, tw * 0.14, tw * 0.15]
    head = [
        Paragraph("<font color='white'><b>Imponible</b></font>", cell_st),
        Paragraph("<font color='white'><b>Apellido y nombre</b></font>", cell_st),
        Paragraph("<font color='white'><b>Recurso</b></font>", cell_st),
        Paragraph("<font color='white'><b>Capital</b></font>", cell_st),
        Paragraph("<font color='white'><b>Intereses</b></font>", cell_st),
        Paragraph("<font color='white'><b>Deuda total</b></font>", cell_st),
    ]
    data = [head]
    for _, row in df_rows.iterrows():
        imp = str(row.get("NRO. IMPONIBLE", "") or "").replace("&", "&amp;").replace("<", "&lt;")
        nom = str(row.get("APELLIDO Y NOMBRE", "") or "").replace("&", "&amp;").replace("<", "&lt;")
        rec = str(row.get("RECURSO", "") or "").replace("&", "&amp;").replace("<", "&lt;")
        data.append(
            [
                Paragraph(imp, cell_st),
                Paragraph(nom, cell_st),
                Paragraph(rec, cell_st),
                fmt_peso(row.get("CAPITAL", 0)),
                fmt_peso(row.get("INTERÉS", 0)),
                fmt_peso(row.get("DEUDA TOTAL", 0)),
            ]
        )
    tbl = Table(data, repeatRows=1, colWidths=col_w)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _PDF_AZUL),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("FONTSIZE", (0, 1), (-1, -1), 7.5),
                ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4F9")]),
                ("GRID", (0, 0), (-1, -1), 0.35, _PDF_AZUL_CLARO),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(tbl)
    doc.build(story, onFirstPage=_pdf_draw_header_footer, onLaterPages=_pdf_draw_header_footer)
    return buf.getvalue()


def merge_datos(existing: pd.DataFrame, incoming: pd.DataFrame) -> pd.DataFrame:
    """
    Por clave imponible + recurso:
    - Coincide con archivo nuevo: actualiza solo montos (capital, interés, deuda).
    - Estaba cargado y NO viene en el nuevo archivo: deuda en $0 (regularizó / ya no figura).
    - Líneas nuevas del archivo: se agregan al final.
    """
    if existing.empty:
        return incoming.copy().reset_index(drop=True)
    if incoming.empty:
        ex = existing.copy()
        for c in ["CAPITAL", "INTERÉS", "MULTA", "DEUDA TOTAL"]:
            if c in ex.columns:
                ex[c] = 0
        return ex.reset_index(drop=True)

    ex = existing.copy().reset_index(drop=True)
    inc = incoming.copy().reset_index(drop=True)
    if "NRO. IMPONIBLE" not in ex.columns or "NRO. IMPONIBLE" not in inc.columns:
        return pd.concat([ex, inc], ignore_index=True)

    ex["_k"] = gestion_row_key(ex)
    inc["_k"] = gestion_row_key(inc)

    debt_cols = [c for c in ["CAPITAL", "INTERÉS", "MULTA", "DEUDA TOTAL"] if c in inc.columns]
    inc_dedup = inc.drop_duplicates("_k", keep="last").set_index("_k")

    ex_keys = set(ex["_k"].tolist())
    inc_keys = set(inc["_k"].tolist())

    for k in ex_keys & inc_keys:
        try:
            row = inc_dedup.loc[k]
            if isinstance(row, pd.DataFrame):
                row = row.iloc[0]
            for c in debt_cols:
                if c in ex.columns:
                    ex.loc[ex["_k"] == k, c] = row[c]
        except Exception:
            continue

    for k in ex_keys - inc_keys:
        for c in debt_cols:
            if c in ex.columns:
                ex.loc[ex["_k"] == k, c] = 0

    new_keys = inc_keys - ex_keys
    if new_keys:
        inc_new = inc[inc["_k"].isin(new_keys)].drop(columns=["_k"], errors="ignore")
        ex = ex.drop(columns=["_k"], errors="ignore")
        ex = pd.concat([ex, inc_new], ignore_index=True)
    else:
        ex = ex.drop(columns=["_k"], errors="ignore")

    return ex


def persist_state():
    payload = {
        "df_datos": st.session_state.df_datos,
        "df_gestion": st.session_state.df_gestion,
        "last_update": st.session_state.last_update,
        "last_file_load": st.session_state.last_file_load,
    }
    pd.to_pickle(payload, STATE_FILE)


def load_persistent_state():
    if st.session_state.get("_persist_loaded"):
        return
    st.session_state["_persist_loaded"] = True

    if not os.path.exists(STATE_FILE):
        return
    try:
        payload = pd.read_pickle(STATE_FILE)
        df_datos = payload.get("df_datos", pd.DataFrame())
        df_gestion = payload.get("df_gestion", pd.DataFrame(columns=GESTION_COLS))

        if isinstance(df_datos, pd.DataFrame):
            st.session_state.df_datos = normalize_numeric_columns(df_datos)
        if isinstance(df_gestion, pd.DataFrame):
            st.session_state.df_gestion = order_gestion_df(df_gestion)
            st.session_state.df_gestion = normalize_numeric_columns(
                st.session_state.df_gestion, ["CAPITAL", "INTERÉS", "DEUDA TOTAL"]
            )

        st.session_state.last_update = payload.get("last_update", st.session_state.last_update)
        st.session_state.last_file_load = payload.get("last_file_load", st.session_state.last_file_load)
    except Exception:
        # Si el archivo está corrupto, no bloqueamos la app.
        pass


# ─── PARSERS ──────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def parse_xlsm(raw_bytes: bytes) -> dict:
    """Parse XLSM/XLSX from R.A.F.A.M. — detects DATOS and GESTION sheets."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    results = {}

    def sheet_to_df(ws, key_cols):
        headers = None
        data = []
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                if row and any(str(c or "").strip().upper() in key_cols for c in row):
                    headers = [str(c or "").strip().upper() for c in row]
                continue
            if any(c is not None for c in row):
                data.append(dict(zip(headers, row)))
        if data:
            return pd.DataFrame(data)
        return pd.DataFrame()

    if "DATOS" in wb.sheetnames:
        df = sheet_to_df(wb["DATOS"], {"RECURSO", "NRO. IMPONIBLE", "APELLIDO Y NOMBRE"})
        if not df.empty:
            results["datos"] = df

    if "GESTION" in wb.sheetnames:
        df = sheet_to_df(wb["GESTION"], {"CUIM", "NRO. IMPONIBLE", "NOTIF. LEGAL"})
        if not df.empty:
            results["gestion"] = df

    wb.close()
    return results


def parse_raw_system_csv(raw_bytes: bytes) -> pd.DataFrame:
    """Parse raw R.A.F.A.M. export CSV (formato nativo con comas)."""
    decoded = raw_bytes.decode("latin-1", errors="replace")
    recurso = ""
    rows = []
    for parts in csv.reader(io.StringIO(decoded), delimiter=",", quotechar='"'):
        if len(parts) >= 11:
            cuim   = parts[2].strip()
            nombre = parts[3].strip()
            if cuim.isdigit() and nombre:
                try:
                    capital = float(parts[6]) if parts[6].strip() else 0
                    interes = float(parts[8]) if parts[8].strip() else 0
                    total   = float(parts[10].strip()) if parts[10].strip() else 0
                    rows.append({
                        "RECURSO":           recurso or "Sin recurso",
                        "TIPO IMPONIBLE":    "INMUEBLE",
                        "NRO. IMPONIBLE":    cuim,
                        "APELLIDO Y NOMBRE": nombre,
                        "CAPITAL":           capital,
                        "INTERÉS":           interes,
                        "MULTA":             0,
                        "DEUDA TOTAL":       total,
                    })
                except Exception:
                    pass
            if parts[0].strip().startswith("Recurso:") and len(parts) >= 2:
                recurso = parts[1].strip().strip('"')
    return pd.DataFrame(rows)


def merge_gestion(df_datos: pd.DataFrame, df_gestion_existing: pd.DataFrame) -> pd.DataFrame:
    """
    Actualiza solo CAPITAL, INTERÉS y DEUDA TOTAL desde el archivo de datos.
    Conserva todo lo demás tal como está en gestión (notificaciones, estado, CUIM, nombre, etc.).
    Si una clave imponible|recurso no viene en el archivo nuevo, los montos pasan a 0.
    """
    if df_datos.empty:
        return order_gestion_df(df_gestion_existing)

    df_new = df_datos.copy()
    df_new["NRO. IMPONIBLE"] = df_new["NRO. IMPONIBLE"].astype(str).str.strip()
    df_new = normalize_numeric_columns(df_new.copy(), ["CAPITAL", "INTERÉS", "DEUDA TOTAL"])
    df_new["_k"] = gestion_row_key(df_new)

    amt_cols = [c for c in ["CAPITAL", "INTERÉS", "DEUDA TOTAL"] if c in df_new.columns]
    if not amt_cols:
        return order_gestion_df(df_gestion_existing)

    lookup = df_new.drop_duplicates("_k", keep="last").set_index("_k")[amt_cols]
    lu = lookup.add_suffix("_sys").reset_index()

    existing = migrate_legacy_gestion(df_gestion_existing.copy())

    if existing.empty:
        deuda_col = pd.to_numeric(df_new.get("DEUDA TOTAL", 0), errors="coerce").fillna(0)
        skeleton = pd.DataFrame({
            "CUIM":              df_new.get("CUIM", pd.Series([""] * len(df_new))),
            "APELLIDO Y NOMBRE": df_new["APELLIDO Y NOMBRE"],
            "NRO. IMPONIBLE":    df_new["NRO. IMPONIBLE"],
            "RECURSO":           df_new.get("RECURSO", pd.Series([""] * len(df_new))),
            "TIPO IMPONIBLE":    df_new.get("TIPO IMPONIBLE", pd.Series(["INMUEBLE"] * len(df_new))),
            "CAPITAL":           pd.to_numeric(df_new.get("CAPITAL", 0), errors="coerce").fillna(0),
            "INTERÉS":           pd.to_numeric(df_new.get("INTERÉS", 0), errors="coerce").fillna(0),
            "DEUDA TOTAL":       deuda_col,
            "TELÉFONO":          "",
            "MAIL":              "",
            "OBSERVACIONES":     "",
            "NOTIF. LEGAL":      pd.Series([pd.NaT] * len(df_new)),
            "NOTIF. CÉDULA":     pd.Series([pd.NaT] * len(df_new)),
            "CARTA DOC.":        pd.Series([pd.NaT] * len(df_new)),
            "ESTADO GESTIÓN":    "Pendiente",
        })
        skeleton["_k"] = gestion_row_key(skeleton)
        return order_gestion_df(skeleton.drop(columns=["_k"], errors="ignore"))

    existing["_k"] = gestion_row_key(existing)
    # No eliminar filas con la misma clave: cada fila conserva notificaciones/estado propios;
    # los montos del sistema se aplican a todas las que comparten imponible|recurso.

    ex2 = existing.merge(lu, on="_k", how="left")
    sys_cols = [f"{c}_sys" for c in amt_cols]
    for c in amt_cols:
        sc = f"{c}_sys"
        if sc in ex2.columns:
            ex2[c] = pd.to_numeric(ex2[sc], errors="coerce").fillna(0)
    ex2.drop(columns=[c for c in sys_cols if c in ex2.columns], inplace=True, errors="ignore")

    ex2 = ex2.drop(columns=["_k"], errors="ignore")

    ex_keys = set(existing["_k"].tolist())
    new_keys = set(lookup.index) - ex_keys
    if new_keys:
        add_src = df_new[df_new["_k"].isin(new_keys)].drop_duplicates("_k", keep="last")
        new_rows = []
        for _, r in add_src.iterrows():
            new_rows.append({
                "CUIM":              "" if pd.isna(r.get("CUIM")) else str(r.get("CUIM", "")),
                "APELLIDO Y NOMBRE": r.get("APELLIDO Y NOMBRE", ""),
                "NRO. IMPONIBLE":    r["NRO. IMPONIBLE"],
                "RECURSO":           "" if pd.isna(r.get("RECURSO")) else str(r.get("RECURSO", "")),
                "TIPO IMPONIBLE":    r.get("TIPO IMPONIBLE", "INMUEBLE") or "INMUEBLE",
                "CAPITAL":           float(pd.to_numeric(r.get("CAPITAL"), errors="coerce") or 0),
                "INTERÉS":           float(pd.to_numeric(r.get("INTERÉS"), errors="coerce") or 0),
                "DEUDA TOTAL":       float(pd.to_numeric(r.get("DEUDA TOTAL"), errors="coerce") or 0),
                "TELÉFONO":          "",
                "MAIL":              "",
                "OBSERVACIONES":     "",
                "NOTIF. LEGAL":      pd.NaT,
                "NOTIF. CÉDULA":     pd.NaT,
                "CARTA DOC.":        pd.NaT,
                "ESTADO GESTIÓN":    "Pendiente",
            })
        ex2 = pd.concat([ex2, pd.DataFrame(new_rows)], ignore_index=True)

    return order_gestion_df(normalize_numeric_columns(ex2, ["CAPITAL", "INTERÉS", "DEUDA TOTAL"]))


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
init_state()

with st.sidebar:
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        c_logo1, c_logo2, c_logo3 = st.columns([1, 2, 1])
        with c_logo2:
            st.image(LOGO_PATH, width=130)
    st.markdown("""
    <div style='text-align:center; padding: 0.2rem 0 0.5rem;'>
        <div style='font-size:0.95rem; font-weight:700; color:#E8F0FB; margin-top:0.3rem;'>
            Gestión de Deuda
        </div>
        <div style='font-size:0.72rem; color:#90B8D8; margin-top:0.1rem;'>
            Municipalidad de Gral. La Madrid
        </div>
    </div>
    <hr style='border-color:rgba(255,255,255,0.15); margin:0.5rem 0 1rem;'>
    """, unsafe_allow_html=True)

    pages = {
        "dashboard":   ("📊", "Dashboard"),
        "importar":    ("📥", "Importar Datos"),
        "gestion":     ("📋", "Gestión de Deuda"),
        "exportar":    ("💾", "Exportar"),
    }

    for key, (icon, label) in pages.items():
        if st.button(f"{icon}  {label}", key=f"nav_{key}", use_container_width=True):
            st.session_state.page = key
            st.rerun()

    st.markdown("<hr style='border-color:rgba(255,255,255,0.15); margin: 1rem 0;'>", unsafe_allow_html=True)

    df_g = st.session_state.df_gestion
    df_d = st.session_state.df_datos
    n_records = len(df_d) if not df_d.empty else 0
    n_gestion = len(df_g) if not df_g.empty else 0
    if not df_g.empty and "ESTADO GESTIÓN" in df_g.columns:
        est = df_g["ESTADO GESTIÓN"].astype(str).str.strip()
        n_pend = int((est == "Pendiente").sum())
        n_eng = int((est == "En Gestión").sum())
        n_leg = int((est == "Legales").sum())
        n_reg = int((est == "Regularizado").sum())
    else:
        n_pend = n_eng = n_leg = n_reg = 0

    st.markdown(f"""
    <div style='font-size:0.75rem; color:#90B8D8; padding: 0 0.5rem;'>
        <div>📦 Registros en base: <b style='color:white'>{n_records:,}</b></div>
        <div style='margin-top:0.3rem;'>📋 Total gestión: <b style='color:white'>{n_gestion:,}</b></div>
        <div style='margin-top:0.3rem;'>⏳ Pendientes: <b style='color:white'>{n_pend:,}</b></div>
        <div style='margin-top:0.3rem;'>🛠️ En gestión: <b style='color:white'>{n_eng:,}</b></div>
        <div style='margin-top:0.3rem;'>⚖️ Legales: <b style='color:white'>{n_leg:,}</b></div>
        <div style='margin-top:0.3rem;'>✅ Regularizados: <b style='color:white'>{n_reg:,}</b></div>
        {"<div style='margin-top:0.3rem; color:#7BAFD4;'>🕐 Actualizado: " + str(st.session_state.last_update) + "</div>" if st.session_state.last_update else ""}
    </div>
    """, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
def page_dashboard():
    st.markdown("""
    <div class='page-banner'>
        <h2>📊 Dashboard de Deuda Municipal</h2>
        <p>Municipalidad de General La Madrid · R.A.F.A.M. · Cuenta Corriente de Imponibles</p>
    </div>
    """, unsafe_allow_html=True)

    df = st.session_state.df_datos
    df_g = st.session_state.df_gestion

    # Fallback al dataset de gestión cuando no hay DATOS (ej. restauración de gestión)
    if df.empty and not df_g.empty:
        df = df_g

    if df.empty:
        st.markdown("""
        <div class='info-box'>
            ⚠️ <b>Sin datos cargados.</b> Ir a <b>📥 Importar Datos</b> para cargar planillas del sistema.
        </div>
        """, unsafe_allow_html=True)
        return

    if st.session_state.last_file_load:
        st.markdown(f"""
        <div class='info-box'>
            🕐 <b>Última carga de archivo:</b> {st.session_state.last_file_load}
        </div>
        """, unsafe_allow_html=True)

    df = normalize_numeric_columns(df.copy(), ["CAPITAL", "INTERÉS", "DEUDA TOTAL"])
    if "DEUDA TOTAL" not in df.columns:
        st.warning("No se encontró la columna DEUDA TOTAL para construir el dashboard.")
        return

    df_pos = df[df["DEUDA TOTAL"] > 0]
    total_deuda     = df_pos["DEUDA TOTAL"].sum()
    total_capital   = df_pos["CAPITAL"].sum() if "CAPITAL" in df_pos.columns else 0
    total_intereses = df_pos["INTERÉS"].sum() if "INTERÉS" in df_pos.columns else 0
    n_imponibles    = len(df_pos)
    pct_int = total_intereses / total_deuda if total_deuda else 0

    # ── KPIs ──
    st.markdown("<div class='sec-header'>RESUMEN FINANCIERO</div>", unsafe_allow_html=True)
    k1, k2, k3, k4 = st.columns(4)

    kpis = [
        (k1, "", "💰 Deuda Total",         fmt_peso(total_deuda),   "Capital + Intereses"),
        (k2, "verde",  "🏛️ Capital",       fmt_peso(total_capital), f"{total_capital/total_deuda*100:.1f}% del total" if total_deuda else "—"),
        (k3, "naranja","📈 Intereses",      fmt_peso(total_intereses), f"{pct_int*100:.1f}% del total"),
        (k4, "rojo",   "📋 Con Deuda",      f"{n_imponibles:,}",     "imponibles activos"),
    ]
    for col, cls, label, value, sub in kpis:
        with col:
            st.markdown(f"""
            <div class='kpi-card {cls}'>
                <div class='kpi-label'>{label}</div>
                <div class='kpi-value {cls}'>{value}</div>
                <div class='kpi-sub'>{sub}</div>
            </div>""", unsafe_allow_html=True)

    # KPIs de gestión (si hay seguimiento cargado)
    if not df_g.empty and "ESTADO GESTIÓN" in df_g.columns:
        st.markdown("<div class='sec-header'>SEGUIMIENTO DE GESTIÓN</div>", unsafe_allow_html=True)
        df_g_aux = normalize_numeric_columns(df_g.copy(), ["DEUDA TOTAL"])
        n_pend = ((df_g_aux.get("ESTADO GESTIÓN", pd.Series([""] * len(df_g_aux))).astype(str) != "Regularizado") &
                  (df_g_aux.get("DEUDA TOTAL", pd.Series([0] * len(df_g_aux))) > 0)).sum()
        n_reg = (df_g_aux.get("ESTADO GESTIÓN", pd.Series([""] * len(df_g_aux))).astype(str) == "Regularizado").sum()
        nl = df_g_aux.get("NOTIF. LEGAL", pd.Series([pd.NaT] * len(df_g_aux)))
        n_notif = sum(notif_tiene_fecha(v) for v in nl)

        g1, g2, g3 = st.columns(3)
        g1.metric("Pendientes", f"{int(n_pend):,}")
        g2.metric("Regularizados", f"{int(n_reg):,}")
        g3.metric("Notif. legal enviadas", f"{int(n_notif):,}")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts ──
    col_a, col_b = st.columns([1.1, 0.9])

    with col_a:
        st.markdown("<div class='sec-header'>DISTRIBUCIÓN CAPITAL vs INTERESES POR RECURSO</div>", unsafe_allow_html=True)
        if "RECURSO" in df_pos.columns:
            df_rec = df_pos.groupby("RECURSO").agg(
                Capital=("CAPITAL", "sum"),
                Intereses=("INTERÉS", "sum"),
            ).reset_index().sort_values("Capital", ascending=True)

            fig = go.Figure()
            fig.add_trace(go.Bar(
                name="Capital", y=df_rec["RECURSO"], x=df_rec["Capital"],
                orientation="h", marker_color="#2E6DA4",
                hovertemplate="<b>%{y}</b><br>Capital: $%{x:,.0f}<extra></extra>"
            ))
            fig.add_trace(go.Bar(
                name="Intereses", y=df_rec["RECURSO"], x=df_rec["Intereses"],
                orientation="h", marker_color="#D68910",
                hovertemplate="<b>%{y}</b><br>Intereses: $%{x:,.0f}<extra></extra>"
            ))
            fig.update_layout(
                barmode="stack", height=320,
                margin=dict(l=10, r=10, t=10, b=10),
                paper_bgcolor="white", plot_bgcolor="white",
                font=dict(family="IBM Plex Sans", size=11),
                legend=dict(orientation="h", y=-0.15),
                xaxis=dict(tickformat="$,.0f", gridcolor="#EEF2F7"),
                yaxis=dict(gridcolor="#EEF2F7"),
            )
            st.plotly_chart(fig, use_container_width=True)

    with col_b:
        st.markdown("<div class='sec-header'>COMPOSICIÓN DE LA DEUDA</div>", unsafe_allow_html=True)
        fig2 = go.Figure(go.Pie(
            labels=["Capital", "Intereses"],
            values=[total_capital, total_intereses],
            hole=0.55,
            marker_colors=["#2E6DA4", "#D68910"],
            textinfo="label+percent",
            hovertemplate="%{label}: $%{value:,.0f}<extra></extra>",
        ))
        fig2.update_layout(
            height=320,
            margin=dict(l=10, r=10, t=10, b=10),
            paper_bgcolor="white",
            font=dict(family="IBM Plex Sans", size=11),
            showlegend=False,
            annotations=[dict(
                text=f"<b>{n_imponibles:,}</b><br>imponibles",
                x=0.5, y=0.5, font_size=14, showarrow=False
            )]
        )
        st.plotly_chart(fig2, use_container_width=True)

    # ── Tabla por recurso ──
    if "RECURSO" in df_pos.columns:
        st.markdown("<div class='sec-header'>DEUDA POR RECURSO</div>", unsafe_allow_html=True)
        df_rec_tbl = df_pos.groupby("RECURSO").agg(
            Imponibles=("NRO. IMPONIBLE", "count"),
            Capital=("CAPITAL", "sum"),
            Intereses=("INTERÉS", "sum"),
            Deuda_Total=("DEUDA TOTAL", "sum")
        ).reset_index()
        df_rec_tbl["% del Total"] = (df_rec_tbl["Deuda_Total"] / total_deuda * 100).round(1)
        df_rec_tbl = df_rec_tbl.sort_values("Deuda_Total", ascending=False)
        df_rec_tbl.columns = ["Recurso", "Imponibles", "Capital ($)", "Intereses ($)", "Deuda Total ($)", "% del Total"]
        for col in ["Capital ($)", "Intereses ($)", "Deuda Total ($)"]:
            df_rec_tbl[col] = df_rec_tbl[col].apply(fmt_peso)
        df_rec_tbl["% del Total"] = df_rec_tbl["% del Total"].apply(lambda x: f"{x:.1f}%")
        st.dataframe(df_rec_tbl, use_container_width=True, hide_index=True,
                     column_config={"Recurso": st.column_config.TextColumn(width="large")})

    # ── Top 50 deudores ──
    st.markdown("<div class='sec-header'>TOP 50 MAYORES DEUDORES</div>", unsafe_allow_html=True)
    cols_top = [c for c in ["NRO. IMPONIBLE", "APELLIDO Y NOMBRE", "RECURSO", "CAPITAL", "INTERÉS", "DEUDA TOTAL"] if c in df_pos.columns]
    if cols_top and "DEUDA TOTAL" in df_pos.columns:
        top10 = df_pos.nlargest(50, "DEUDA TOTAL")[cols_top].copy()
        top10.insert(0, "#", range(1, len(top10) + 1))
        for col in ["CAPITAL", "INTERÉS", "DEUDA TOTAL"]:
            if col in top10.columns:
                top10[col] = top10[col].apply(fmt_peso)
        st.dataframe(top10, use_container_width=True, hide_index=True)
    else:
        st.info("No hay columnas suficientes para mostrar el ranking de deudores.")


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: IMPORTAR
# ═══════════════════════════════════════════════════════════════════════════════
def page_importar():
    st.markdown("""
    <div class='page-banner'>
        <h2>📥 Importar Datos del Sistema</h2>
        <p>Cargá planillas del R.A.F.A.M. en formato XLS/XLSX/XLSM o el archivo de gestión guardado</p>
    </div>
    """, unsafe_allow_html=True)

    if st.session_state.get("_import_pending"):
        p = st.session_state.pop("_import_pending")
        st.session_state["_import_job"] = {
            "kind": p.get("kind"),
            "raw": p.get("raw"),
            "step": 0,
        }
        st.rerun()

    job = st.session_state.get("_import_job")
    if job is not None:
        st.markdown("### ⏳ Importando archivo")
        st.info(
            "La barra avanza en varios pasos (Streamlit vuelve a cargar la página entre cada uno). "
            "**Solo se actualizan montos** (capital, intereses, deuda); notificaciones, observaciones, "
            "estado y el resto **no se borran**."
        )
        status = st.empty()
        try:
            progress_ui = st.progress(0, text="…")
        except TypeError:
            progress_ui = st.progress(0)

        def _prog(pct: int, msg: str) -> None:
            p = min(100, max(0, int(pct)))
            status.markdown(f"**{msg}**")
            try:
                progress_ui.progress(p / 100.0, text=msg)
            except TypeError:
                progress_ui.progress(p / 100.0)

        kind = job.get("kind")
        raw = job.get("raw")
        step = int(job.get("step", 0))

        try:
            if not raw:
                st.error("No hay datos de archivo para importar.")
                st.session_state["_import_job"] = None
            elif kind == "xlsx":
                if step == 0:
                    _prog(5, "Preparando importación…")
                    job["step"] = 1
                    st.rerun()
                elif step == 1:
                    _prog(22, "Leyendo planilla (hoja DATOS)…")
                    results = parse_xlsm(raw)
                    if "datos" not in results or results["datos"].empty:
                        st.error("No se encontró la hoja DATOS en el archivo.")
                        st.session_state["_import_job"] = None
                    else:
                        job["_xlsx_results"] = results
                        job["step"] = 2
                        st.rerun()
                elif step == 2:
                    _prog(48, "Fusionando cuenta corriente (solo montos)…")
                    results = job.get("_xlsx_results") or {}
                    df_d = normalize_numeric_columns(results["datos"])
                    merged_datos = merge_datos(st.session_state.df_datos, df_d)
                    st.session_state.df_datos = normalize_numeric_columns(merged_datos)
                    job["step"] = 3
                    st.rerun()
                elif step == 3:
                    _prog(
                        72,
                        "Actualizando CAPITAL / INTERÉS / DEUDA en gestión — lo demás no se toca…",
                    )
                    df_g = merge_gestion(st.session_state.df_datos, st.session_state.df_gestion)
                    st.session_state.df_gestion = df_g
                    job["step"] = 4
                    st.rerun()
                elif step == 4:
                    _prog(90, "Guardando en disco…")
                    st.session_state.last_update = datetime.now().strftime("%d/%m/%Y %H:%M")
                    st.session_state.last_file_load = st.session_state.last_update
                    persist_state()
                    _prog(100, "Listo.")
                    st.success(
                        "Importación terminada: se actualizaron solo los montos; lo cargado a mano se conservó."
                    )
                    st.session_state["_import_job"] = None
                    st.rerun()
            elif kind == "csv":
                if step == 0:
                    _prog(5, "Preparando importación…")
                    job["step"] = 1
                    st.rerun()
                elif step == 1:
                    _prog(28, "Leyendo CSV…")
                    df_csv = normalize_numeric_columns(parse_raw_system_csv(raw))
                    if df_csv.empty:
                        st.error("No se pudieron leer registros del CSV.")
                        st.session_state["_import_job"] = None
                    else:
                        job["_csv_df"] = df_csv
                        job["step"] = 2
                        st.rerun()
                elif step == 2:
                    _prog(52, "Fusionando cuenta corriente…")
                    df_csv = job["_csv_df"]
                    merged_datos = merge_datos(st.session_state.df_datos, df_csv)
                    st.session_state.df_datos = normalize_numeric_columns(merged_datos)
                    job["step"] = 3
                    st.rerun()
                elif step == 3:
                    _prog(75, "Actualizando solo montos en gestión…")
                    df_g = merge_gestion(st.session_state.df_datos, st.session_state.df_gestion)
                    st.session_state.df_gestion = df_g
                    job["step"] = 4
                    st.rerun()
                elif step == 4:
                    _prog(90, "Guardando en disco…")
                    st.session_state.last_update = datetime.now().strftime("%d/%m/%Y %H:%M")
                    st.session_state.last_file_load = st.session_state.last_update
                    persist_state()
                    _prog(100, "Listo.")
                    st.success(
                        "Importación terminada: se actualizaron solo los montos; lo cargado a mano se conservó."
                    )
                    st.session_state["_import_job"] = None
                    st.rerun()
            else:
                st.error("Tipo de importación no reconocido.")
                st.session_state["_import_job"] = None
        except Exception as e:
            st.error(f"Error al importar: {e}")
            st.session_state["_import_job"] = None

        if st.session_state.get("_import_job") is not None:
            return

    tab1, tab2, tab3 = st.tabs([
        "📄 Planilla del Sistema (.xls/.xlsx/.xlsm)",
        "💾 Archivo de Gestión (.xlsx)",
        "📋 CSV del sistema",
    ])

    # ── TAB 1: Planilla del sistema ──
    with tab1:
        st.markdown("""
        <div class='info-box'>
            Subí la planilla descargada del R.A.F.A.M. con extensión <b>.xls</b>, <b>.xlsx</b> o <b>.xlsm</b>.
            Se detectan automáticamente las hojas <b>DATOS</b> y <b>GESTION</b>.
            <br><br>
            <b>Importante:</b> al confirmar, solo se actualizan <b>CAPITAL, INTERÉS y DEUDA TOTAL</b> desde la hoja DATOS.
            Lo que cargaste a mano (notificaciones, estado, teléfono, observaciones, etc.) <b>se conserva</b>.
            La hoja GESTION del archivo <b>no se usa</b> al importar, para no pisar tu trabajo en la app.
        </div>
        """, unsafe_allow_html=True)

        uploaded = st.file_uploader(
            "Seleccioná el archivo de planilla",
            type=["xls", "xlsx", "xlsm"],
            key="uploader_xlsx",
        )

        if uploaded:
            with st.spinner("Procesando archivo..."):
                raw = uploaded.read()
                try:
                    results = parse_xlsm(raw)
                except Exception as e:
                    st.error(f"Error al leer el archivo: {e}")
                    return
            file_bytes = uploaded.getvalue()

            if "datos" not in results or results["datos"].empty:
                try:
                    df_test = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, engine="openpyxl")
                    df_test.columns = df_test.columns.str.strip().str.upper()

                    if "DEUDA TOTAL" in df_test.columns and "NRO. IMPONIBLE" in df_test.columns:
                        df_test = normalize_numeric_columns(df_test)
                        merged_d = merge_datos(st.session_state.df_datos, df_test)
                        st.session_state.df_datos = normalize_numeric_columns(merged_d)
                        st.session_state.df_gestion = merge_gestion(
                            st.session_state.df_datos, st.session_state.df_gestion
                        )
                        st.session_state.last_update = datetime.now().strftime("%d/%m/%Y %H:%M")
                        st.session_state.last_file_load = st.session_state.last_update
                        persist_state()
                        st.success(
                            f"✅ {len(df_test):,} filas leídas de la primera hoja — "
                            "montos fusionados sin borrar tu gestión manual."
                        )
                        st.rerun()
                    else:
                        st.warning("⚠️ No se encontró la hoja DATOS ni formato de gestión reconocido.")

                except Exception as e:
                    st.error(f"Error al leer archivo: {e}")
                # Mostrar hojas disponibles si es posible
                try:
                    from openpyxl import load_workbook
                    wb2 = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    st.info(f"Hojas encontradas en el archivo: {', '.join(wb2.sheetnames)}")
                    wb2.close()
                except Exception:
                    pass
                return

            df_d = results["datos"]
            df_d = normalize_numeric_columns(df_d)

            df_pos = df_d[df_d.get("DEUDA TOTAL", pd.Series([0] * len(df_d))) > 0] if "DEUDA TOTAL" in df_d.columns else df_d

            st.success(f"✅ **{len(df_d):,}** registros encontrados en hoja DATOS")

            c1, c2, c3 = st.columns(3)
            with c1: st.metric("Imponibles con deuda", f"{len(df_pos):,}")
            with c2: st.metric("Deuda total", fmt_peso(df_d["DEUDA TOTAL"].sum()) if "DEUDA TOTAL" in df_d.columns else "—")
            with c3: st.metric("Recursos", df_d["RECURSO"].nunique() if "RECURSO" in df_d.columns else "—")

            with st.expander("Vista previa de datos (primeras 5 filas)"):
                st.dataframe(df_d.head(5), hide_index=True, use_container_width=True)

            if "gestion" in results and not results["gestion"].empty:
                st.caption(
                    f"ℹ️ El archivo incluye hoja GESTION ({len(results['gestion']):,} filas). "
                    "Al confirmar **no** se usa para no pisar datos cargados en la app; solo cuenta la hoja DATOS."
                )

            if st.button("✅ Confirmar importación", use_container_width=True, key="confirm_xlsx"):
                st.session_state["_import_pending"] = {"kind": "xlsx", "raw": raw}
                st.rerun()

    # ── TAB 2: Archivo de gestión guardado ──
    with tab2:
        st.markdown("""
        <div class='info-box'>
            Restaurá un archivo de gestión exportado previamente desde esta aplicación (.xlsx).
            Los datos de seguimiento (notificaciones, estados, contactos) se restaurarán tal como fueron guardados.
        </div>
        """, unsafe_allow_html=True)

        uploaded_g = st.file_uploader(
            "Archivo de gestión (.xlsx)",
            type=["xlsx"],
            key="uploader_gestion",
        )

        if uploaded_g:
            try:
                df_restore = pd.read_excel(io.BytesIO(uploaded_g.read()), sheet_name=0)
                df_restore.columns = df_restore.columns.str.strip().str.upper()

# asegurar columnas mínimas
                for col in GESTION_COLS:
                    if col not in df_restore.columns:
                        df_restore[col] = ""

                df_restore = df_restore[GESTION_COLS]
                df_restore = normalize_numeric_columns(df_restore, ["CAPITAL", "INTERÉS", "DEUDA TOTAL"])

                st.success(f"✅ {len(df_restore):,} registros encontrados")
                st.dataframe(df_restore.head(5), hide_index=True)

                if st.button("Restaurar gestión", use_container_width=True):
                    st.session_state.df_gestion = df_restore
                    st.session_state.last_update = datetime.now().strftime("%d/%m/%Y %H:%M")
                    st.session_state.last_file_load = st.session_state.last_update
                    persist_state()
                    st.success("Gestión restaurada.")
                    st.rerun()

            except Exception as e:
                st.error(f"Error al leer el archivo: {e}")

    # ── TAB 3: CSV del sistema ──
    with tab3:
        st.markdown("""
        <div class='info-box'>
            Subí el CSV/XLS original descargado del sistema R.A.F.A.M. (formato nativo con columnas separadas por comas).
        </div>
        """, unsafe_allow_html=True)

        uploaded_csv = st.file_uploader(
            "CSV del sistema",
            type=["csv", "xls", "txt"],
            key="uploader_csv",
        )

        if uploaded_csv:
            raw = uploaded_csv.read()
            df_csv = normalize_numeric_columns(parse_raw_system_csv(raw))

            if df_csv.empty:
                st.warning("⚠️ No se pudieron detectar registros válidos. Verificá el formato del archivo.")
                return

            df_pos = df_csv[df_csv["DEUDA TOTAL"] > 0] if "DEUDA TOTAL" in df_csv.columns else df_csv
            st.success(f"✅ {len(df_csv):,} registros detectados — {len(df_pos):,} con deuda")

            with st.expander("Vista previa (primeras 10 filas)"):
                st.dataframe(df_csv.head(10), hide_index=True, use_container_width=True)

            if st.button("Importar CSV", use_container_width=True, key="import_csv"):
                st.session_state["_import_pending"] = {"kind": "csv", "raw": raw}
                st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: GESTIÓN
# ═══════════════════════════════════════════════════════════════════════════════
def page_gestion():
    st.markdown("""
    <div class='page-banner'>
        <h2>📋 Gestión de Deuda</h2>
        <p>Búsqueda, seguimiento y notificaciones por contribuyente</p>
    </div>
    """, unsafe_allow_html=True)

    df = order_gestion_df(normalize_numeric_columns(st.session_state.df_gestion.copy()))

    if df.empty:
        st.markdown("""
        <div class='info-box'>
            ⚠️ Sin datos. Importá una planilla primero desde <b>📥 Importar Datos</b>.
        </div>
        """, unsafe_allow_html=True)
        return

    st.markdown("<div class='sec-header'>🔍 FILTROS DE BÚSQUEDA</div>", unsafe_allow_html=True)
    try:
        filt_wrap = st.container(border=True)
    except TypeError:
        filt_wrap = st.container()
    with filt_wrap:
        buscar = st.text_input(
            "Buscar contribuyente (CUIM, imponible o nombre)",
            placeholder="Ej: 12345 o MANSUR o parte del nombre…",
            key="gest_buscar",
        )
        fc1, fc2, fc3, fc4 = st.columns([2, 2, 2, 1.5])
        with fc1:
            search_nombre = st.text_input("Nombre / Apellido", placeholder="Ej: GARCIA, MARIO...", key="gest_nom")
        with fc2:
            search_cuim = st.text_input("CUIM o Nro. Imponible", placeholder="Ej: 1234", key="gest_cuim")
        with fc3:
            recursos_list = ["Todos"] + sorted(df["RECURSO"].dropna().unique().tolist()) if "RECURSO" in df.columns else ["Todos"]
            filtro_recurso = st.selectbox("Recurso", recursos_list, key="gest_rec")
        with fc4:
            filtro_notif = st.selectbox(
                "Notif. legal",
                ["Todas", "Con fecha (notificado)", "Sin fecha (pendiente)"],
                key="gest_not",
            )

        fc5, fc6, _ = st.columns([2, 2, 2])
        with fc5:
            estados_list = ["Todos"] + sorted(df["ESTADO GESTIÓN"].dropna().unique().tolist()) if "ESTADO GESTIÓN" in df.columns else ["Todos"]
            filtro_estado = st.selectbox("Estado de Gestión", estados_list, key="gest_est")
        with fc6:
            solo_con_deuda = st.checkbox("Solo con deuda > 0", value=True, key="gest_solo")

    mask = pd.Series([True] * len(df), index=df.index)
    if buscar and str(buscar).strip():
        q = str(buscar).strip().upper()
        cuim_col = df.get("CUIM", pd.Series([""] * len(df), index=df.index)).astype(str)
        mask &= (
            df["APELLIDO Y NOMBRE"].astype(str).str.upper().str.contains(q, na=False)
            | df["NRO. IMPONIBLE"].astype(str).str.contains(q, na=False)
            | cuim_col.str.upper().str.contains(q, na=False)
        )
    if search_nombre:
        nombre_upper = df["APELLIDO Y NOMBRE"].astype(str).str.upper()
        mask &= nombre_upper.str.contains(search_nombre.upper(), na=False)
    if search_cuim:
        cuim_col = df.get("CUIM", pd.Series([""] * len(df), index=df.index)).astype(str)
        mask &= (
            df["NRO. IMPONIBLE"].astype(str).str.contains(search_cuim, na=False)
            | cuim_col.str.contains(search_cuim, na=False)
        )
    if filtro_recurso != "Todos" and "RECURSO" in df.columns:
        mask &= df["RECURSO"] == filtro_recurso
    if filtro_estado != "Todos" and "ESTADO GESTIÓN" in df.columns:
        mask &= df["ESTADO GESTIÓN"] == filtro_estado
    if solo_con_deuda:
        mask &= df["DEUDA TOTAL"] > 0
    if filtro_notif != "Todas" and "NOTIF. LEGAL" in df.columns:
        nl = df["NOTIF. LEGAL"]
        if filtro_notif == "Con fecha (notificado)":
            mask &= nl.apply(notif_tiene_fecha)
        else:
            mask &= ~nl.apply(notif_tiene_fecha)

    df_filtered = df[mask].copy()
    n_filtered = len(df_filtered)
    total_filt = df_filtered["DEUDA TOTAL"].sum()

    st.markdown(f"""
    <div style='display:flex; gap:1.5rem; margin:0.5rem 0 1rem; align-items:center;'>
        <div style='font-size:0.85rem; color:#475569;'>
            📊 <b>{n_filtered:,}</b> registros · Deuda: <b>{fmt_peso(total_filt)}</b>
        </div>
    </div>
    """, unsafe_allow_html=True)
    est_f = df_filtered.get("ESTADO GESTIÓN", pd.Series([""] * len(df_filtered), index=df_filtered.index)).astype(str).str.strip()
    notificados = sum(notif_tiene_fecha(v) for v in df_filtered.get("NOTIF. LEGAL", pd.Series([pd.NaT] * len(df_filtered), index=df_filtered.index)))
    por_notificar = int(len(df_filtered) - notificados)
    mk1, mk2, mk3, mk4, mk5 = st.columns(5)
    minis = [
        ("Pendientes", int((est_f == "Pendiente").sum())),
        ("En gestión", int((est_f == "En Gestión").sum())),
        ("Legales", int((est_f == "Legales").sum())),
        ("Regularizados", int((est_f == "Regularizado").sum())),
        ("Sin notif. legal", por_notificar),
    ]
    for col, (lbl, val) in zip([mk1, mk2, mk3, mk4, mk5], minis):
        with col:
            st.markdown(
                f"<div class='mini-kpi'><div class='label'>{lbl}</div><div class='value'>{val:,}</div></div>",
                unsafe_allow_html=True,
            )

    st.markdown("<div class='sec-header'>📊 TABLA DE REGISTROS — edición directa</div>", unsafe_allow_html=True)
    st.caption(
        "Marcá **Incluir en suma** en las filas que quieras: abajo se muestra la suma de Capital e Intereses. "
        "**Guardar cambios** aplica ediciones y recalcula Deuda total = Capital + Intereses."
    )

    if df_filtered.empty:
        st.info("No hay registros con estos filtros.")
        return

    df_edit = df_filtered.copy()
    for c in DATE_NOTIF_COLS:
        if c in df_edit.columns:
            df_edit[c] = pd.to_datetime(df_edit[c], errors="coerce")
    df_edit.insert(0, "Incluir en suma", False)
    col_order = ["Incluir en suma"] + GESTION_COLS
    df_edit = df_edit[[c for c in col_order if c in df_edit.columns]]

    cc = {
        "Incluir en suma": st.column_config.CheckboxColumn(
            "Incluir en suma", help="Marcá filas para sumar capital e intereses"
        ),
        "CUIM": st.column_config.TextColumn("CUIM", width="small"),
        "CAPITAL": st.column_config.NumberColumn("Capital", format="%.2f"),
        "INTERÉS": st.column_config.NumberColumn("Intereses", format="%.2f"),
        "DEUDA TOTAL": st.column_config.NumberColumn("Deuda total", format="%.2f", disabled=True),
        "NOTIF. LEGAL": st.column_config.DateColumn("Notif. legal", format="DD/MM/YYYY"),
        "NOTIF. CÉDULA": st.column_config.DateColumn("Notif. cédula", format="DD/MM/YYYY"),
        "CARTA DOC.": st.column_config.DateColumn("Carta doc.", format="DD/MM/YYYY"),
        "ESTADO GESTIÓN": st.column_config.SelectboxColumn("Estado", options=ESTADOS, required=True),
        "OBSERVACIONES": st.column_config.TextColumn("Observaciones", width="large"),
    }

    edited = st.data_editor(
        df_edit,
        use_container_width=True,
        hide_index=True,
        height=520,
        column_config=cc,
        disabled=["DEUDA TOTAL"],
        key="gestion_data_editor",
    )

    sel = edited[edited["Incluir en suma"].fillna(False)]
    s_cap = pd.to_numeric(sel["CAPITAL"], errors="coerce").fillna(0).sum()
    s_int = pd.to_numeric(sel["INTERÉS"], errors="coerce").fillna(0).sum()
    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("Suma Capital (selección)", fmt_peso(s_cap))
    with m2:
        st.metric("Suma Intereses (selección)", fmt_peso(s_int))
    with m3:
        st.metric("Suma total (selección)", fmt_peso(s_cap + s_int))

    if st.button("💾 Guardar cambios de la tabla", type="primary", use_container_width=True, key="gest_save_tbl"):
        for orig_idx, (_, r) in zip(df_filtered.index, edited.iterrows()):
            for c in GESTION_COLS:
                if c in ("CAPITAL", "INTERÉS", "DEUDA TOTAL"):
                    continue
                val = r.get(c)
                if c in DATE_NOTIF_COLS:
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        val = pd.NaT
                    else:
                        val = pd.to_datetime(val, errors="coerce")
                st.session_state.df_gestion.at[orig_idx, c] = val
            cap = float(pd.to_numeric(r.get("CAPITAL"), errors="coerce") or 0)
            inter = float(pd.to_numeric(r.get("INTERÉS"), errors="coerce") or 0)
            st.session_state.df_gestion.at[orig_idx, "CAPITAL"] = cap
            st.session_state.df_gestion.at[orig_idx, "INTERÉS"] = inter
            st.session_state.df_gestion.at[orig_idx, "DEUDA TOTAL"] = cap + inter
        st.session_state.df_gestion = order_gestion_df(normalize_numeric_columns(st.session_state.df_gestion.copy()))
        st.session_state.last_update = datetime.now().strftime("%d/%m/%Y %H:%M")
        persist_state()
        st.success("Cambios guardados.")
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE: EXPORTAR
# ═══════════════════════════════════════════════════════════════════════════════
def page_exportar():
    st.markdown("""
    <div class='page-banner'>
        <h2>💾 Exportar</h2>
        <p>Elegí <strong>PDF</strong> o <strong>Excel</strong> y qué datos incluir — sin vista previa del texto del informe.</p>
    </div>
    """, unsafe_allow_html=True)

    df_g = order_gestion_df(normalize_numeric_columns(st.session_state.df_gestion.copy()))
    df_d = st.session_state.df_datos

    if df_g.empty and df_d.empty:
        st.info("Sin datos para exportar.")
        return

    try:
        opt_wrap = st.container(border=True)
    except TypeError:
        opt_wrap = st.container()
    with opt_wrap:
        formato = st.radio(
            "Formato de salida",
            ["Excel (.xlsx)", "PDF (.pdf)"],
            horizontal=True,
            key="exp_formato",
        )

        fecha = datetime.now().strftime("%Y%m%d_%H%M")

        if formato.startswith("Excel"):
            st.markdown("**Incluir en el archivo**")
            cxa, cxb = st.columns(2)
            with cxa:
                inc_gestion = st.checkbox(
                    "Hoja **GESTION** (planilla municipal)",
                    value=not df_g.empty,
                    disabled=df_g.empty,
                    key="exp_inc_gestion",
                )
            with cxb:
                inc_datos = st.checkbox(
                    "Hoja **DATOS** (archivo importado del sistema)",
                    value=not df_d.empty,
                    disabled=df_d.empty,
                    key="exp_inc_datos",
                )
            if not inc_gestion and not inc_datos:
                st.warning("Marcá al menos una opción para exportar.")
            else:
                try:
                    xbytes = excel_bytes_gestion_profesional(
                        df_g if not df_g.empty else None,
                        df_d if not df_d.empty else None,
                        include_gestion=inc_gestion,
                        include_datos=inc_datos,
                    )
                except ValueError as e:
                    st.error(str(e))
                else:
                    st.download_button(
                        "⬇️ Descargar Excel",
                        data=xbytes,
                        file_name=f"GestionDeuda_{fecha}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_excel",
                    )
        else:
            if df_g.empty:
                st.warning("No hay gestión cargada para armar el PDF.")
            else:
                st.markdown("**Qué registros van al PDF** (tabla imponible, nombre, recurso, montos)")
                rec_opts = (
                    ["Todos los recursos"]
                    + sorted(df_g["RECURSO"].dropna().astype(str).unique().tolist())
                    if "RECURSO" in df_g.columns
                    else ["Todos los recursos"]
                )
                n_max = max(1, len(df_g))
                cant = st.number_input(
                    "Cantidad de filas (ordenadas por deuda total, mayor primero)",
                    min_value=1,
                    max_value=n_max,
                    value=min(50, n_max),
                    key="exp_pdf_n",
                )
                rec_sel = st.selectbox("Recurso", rec_opts, key="exp_pdf_rec")
                solo_deuda = st.checkbox(
                    "Solo imponibles con deuda total mayor a cero",
                    value=True,
                    key="exp_pdf_solo_deuda",
                )
                if not HAS_REPORTLAB:
                    st.warning("Instalá **reportlab**: `pip install reportlab`")
                else:
                    base = df_g.copy()
                    if solo_deuda and "DEUDA TOTAL" in base.columns:
                        base = base[pd.to_numeric(base["DEUDA TOTAL"], errors="coerce").fillna(0) > 0]
                    if rec_sel != "Todos los recursos" and "RECURSO" in base.columns:
                        base = base[base["RECURSO"].astype(str) == rec_sel]
                    base = base.sort_values(
                        "DEUDA TOTAL", ascending=False
                    ).head(int(cant)) if "DEUDA TOTAL" in base.columns else base.head(int(cant))
                    sub = f"Reporte de deuda — {rec_sel} — {len(base)} registros"
                    if base.empty:
                        st.info("No hay filas con los filtros elegidos. Ajustá cantidad, recurso o desmarcá «solo con deuda».")
                    else:
                        pdf_bytes = pdf_bytes_reporte_deuda(base, sub)
                        st.download_button(
                            "⬇️ Descargar PDF",
                            data=pdf_bytes,
                            file_name=f"Informe_de_Deuda_{fecha}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key="dl_pdf",
                        )

    st.markdown("<div class='sec-header'>ESTADÍSTICAS DE GESTIÓN</div>", unsafe_allow_html=True)
    if not df_g.empty:
        df_g2 = df_g.copy()
        if "ESTADO GESTIÓN" in df_g2.columns:
            ec1, ec2 = st.columns(2)
            with ec1:
                df_est = df_g2.groupby("ESTADO GESTIÓN").agg(
                    Cantidad=("NRO. IMPONIBLE", "count"),
                    Deuda=("DEUDA TOTAL", "sum"),
                ).reset_index()
                fig_e = px.bar(
                    df_est, x="ESTADO GESTIÓN", y="Cantidad",
                    color="ESTADO GESTIÓN", text="Cantidad",
                    color_discrete_sequence=px.colors.qualitative.Set2,
                )
                fig_e.update_layout(
                    height=300, showlegend=False,
                    paper_bgcolor="white", plot_bgcolor="white",
                    margin=dict(l=10, r=10, t=10, b=10),
                    font=dict(family="IBM Plex Sans", size=11),
                )
                st.plotly_chart(fig_e, use_container_width=True)
            with ec2:
                def _nnotif(serie):
                    return sum(notif_tiene_fecha(v) for v in serie) if len(serie) else 0

                df_notif = pd.DataFrame({
                    "Tipo": ["Legal", "Cédula", "Carta Doc."],
                    "Notificados": [
                        _nnotif(df_g2.get("NOTIF. LEGAL", pd.Series(dtype=object))),
                        _nnotif(df_g2.get("NOTIF. CÉDULA", pd.Series(dtype=object))),
                        _nnotif(df_g2.get("CARTA DOC.", pd.Series(dtype=object))),
                    ],
                })
                fig_n = px.bar(
                    df_notif, x="Tipo", y="Notificados",
                    text="Notificados", color="Tipo",
                    color_discrete_sequence=["#1E3A5F", "#2E6DA4", "#1A6B3C"],
                )
                fig_n.update_layout(
                    height=300, showlegend=False,
                    paper_bgcolor="white", plot_bgcolor="white",
                    margin=dict(l=10, r=10, t=10, b=10),
                    font=dict(family="IBM Plex Sans", size=11),
                )
                st.plotly_chart(fig_n, use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTER
# ═══════════════════════════════════════════════════════════════════════════════
_routes = {
    "dashboard":   page_dashboard,
    "importar":    page_importar,
    "gestion":     page_gestion,
    "exportar":    page_exportar,
}

_routes.get(st.session_state.page, page_dashboard)()
